import pandas as pd
import tkinter as tk
import tkinter.messagebox as mb
from tkinter import simpledialog, filedialog
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side

def upload_files():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    filetypes = [("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")]

    file1_path = filedialog.askopenfilename(title="Select the first file (SCANNED DATA)", filetypes=filetypes)
    file2_path = filedialog.askopenfilename(title="Select the second file (IP DATA)", filetypes=filetypes)
    output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Save output Excel file")
    
    last_column = simpledialog.askstring("Input", "Enter the last column to compare (e.g., A150):", parent=root)
    
    include_qpseries = simpledialog.askstring("Input", "Do you want to include the QPSERIES column in the comparison? (y/n):", parent=root).lower()

    if file1_path and file2_path and output_path and last_column:
        if check_rollno_columns(file1_path, file2_path, root):
            return
        
        duplicates_found = check_duplicates(file1_path, root) or check_duplicates(file2_path, root)
        suffix_found = check_suffix_in_rollnos(file1_path, file2_path, root)
        
        if duplicates_found or suffix_found:
            proceed = simpledialog.askstring("Proceed", "Duplicates or suffixes found in ROLLNO column. Do you want to proceed? (y/n):", parent=root).lower()
            if proceed == 'n':
                root.destroy()
                return
        
        check_discrepancy(file1_path, file2_path, output_path, last_column, include_qpseries, root)
        mb.showinfo("Success", "Data exported successfully to the chosen path!", parent=root)
    root.destroy()

def check_rollno_columns(file1_path, file2_path, root):
    try:
        df1 = pd.read_excel(file1_path) if file1_path.endswith('.xlsx') else pd.read_csv(file1_path)
        df2 = pd.read_excel(file2_path) if file2_path.endswith('.xlsx') else pd.read_csv(file2_path)
        
        if 'ROLLNO' not in df1.columns or 'ROLLNO' not in df2.columns:
            mb.showerror("Error", "ROLLNO column missing in one or both files.", parent=root)
            return True
        
        # Strip whitespace and convert to string for consistent comparison
        df1['ROLLNO'] = df1['ROLLNO'].astype(str).str.strip()
        df2['ROLLNO'] = df2['ROLLNO'].astype(str).str.strip()

        # Use sets to check if the ROLLNO values are the same
        set1 = set(df1['ROLLNO'])
        set2 = set(df2['ROLLNO'])
        
        if set1 != set2:
            mb.showerror("Error", "The 'ROLLNO' columns do not match between the two files.", parent=root)
            return True
    except Exception as e:
        mb.showerror("Error", f"Error comparing ROLLNO columns: {e}", parent=root)
        return True
    return False

def check_duplicates(file_path, root):
    try:
        df = pd.read_excel(file_path) if file_path.endswith('.xlsx') else pd.read_csv(file_path)
        if 'ROLLNO' not in df.columns:
            mb.showerror("Error", "ROLLNO column missing in the file.", parent=root)
            return True
        duplicates = df[df.duplicated('ROLLNO', keep=False)]
        if not duplicates.empty:
            duplicate_details = duplicates[['ROLLNO']].reset_index().to_string(index=False)
            mb.showinfo("Duplicates Found", f"Duplicate ROLLNO values found:\n\n{duplicate_details}", parent=root)
            return True
    except Exception as e:
        mb.showerror("Error", f"Error reading file: {e}", parent=root)
        return True
    return False

def check_suffix_in_rollnos(file1_path, file2_path, root):
    suffix_pattern = re.compile(r'[^\d]+')

    try:
        df1 = pd.read_excel(file1_path) if file1_path.endswith('.xlsx') else pd.read_csv(file1_path)
        df2 = pd.read_excel(file2_path) if file2_path.endswith('.xlsx') else pd.read_csv(file2_path)
        if 'ROLLNO' not in df1.columns or 'ROLLNO' not in df2.columns:
            mb.showerror("Error", "ROLLNO column missing in one or both files.", parent=root)
            return True
        df1['ROLLNO'] = df1['ROLLNO'].astype(str)
        df2['ROLLNO'] = df2['ROLLNO'].astype(str)
        suffixes_file1 = [suffix_pattern.search(x).group() if suffix_pattern.search(x) else '' for x in df1['ROLLNO']]
        suffixes_file2 = [suffix_pattern.search(x).group() if suffix_pattern.search(x) else '' for x in df2['ROLLNO']]
        unique_suffixes_file1 = list(set(suffixes_file1))
        unique_suffixes_file2 = list(set(suffixes_file2))
        suffixes_file1_str = ', '.join(unique_suffixes_file1)
        suffixes_file2_str = ', '.join(unique_suffixes_file2)
        if unique_suffixes_file1 or unique_suffixes_file2:
            mb.showinfo("Suffixes Found", f"Suffixes found in ROLLNO column:\n\nFile 1: {suffixes_file1_str}\nFile 2: {suffixes_file2_str}", parent=root)
            return True
    except Exception as e:
        mb.showerror("Error", f"Error checking suffixes: {e}", parent=root)
        return True
    return False

def check_discrepancy(file1_path, file2_path, output_path, last_column, include_qpseries, root):
    df1 = pd.read_excel(file1_path) if file1_path.endswith('.xlsx') else pd.read_csv(file1_path)
    df2 = pd.read_excel(file2_path) if file2_path.endswith('.xlsx') else pd.read_csv(file2_path)

    df1.columns = df1.columns.str.strip()
    df2.columns = df2.columns.str.strip()

    df1['ROLLNO'] = df1['ROLLNO'].astype(str)
    df2['ROLLNO'] = df2['ROLLNO'].astype(str)

    df1.sort_values(by='ROLLNO', inplace=True)
    df2.sort_values(by='ROLLNO', inplace=True)

    try:
        last_column_index = int(last_column[1:])
        if last_column_index < 1 or last_column_index > 150:
            raise ValueError
        columns_to_compare = ['ROLLNO'] + [f'A{i}' for i in range(1, last_column_index + 1)]
        if include_qpseries == 'y' and 'QPSERIES' in df1.columns and 'QPSERIES' in df2.columns:
            columns_to_compare.insert(1, 'QPSERIES')
    except ValueError:
        mb.showerror("Error", "Invalid column input. Please enter a valid column like A5 or A150.", parent=root)
        return

    if 'Front side Image' not in df1.columns:
        mb.showerror("Error", "'Front Side Image' column not found in the first file.", parent=root)
        print("Available columns in file 1:", df1.columns)
        return

    df1 = df1[columns_to_compare + ['Front side Image']]
    df2 = df2[columns_to_compare]

    discrepancies = []
    qpseries_rows = []

    for idx in range(len(df1)):
        row1 = df1.iloc[idx]
        row2 = df2.iloc[idx]
        rollno = row1['ROLLNO']
        image_path = row1.get('Front side Image', '')

        for col in columns_to_compare[1:]:
            if row1[col] != row2[col]:
                if col == 'QPSERIES':
                    qpseries_rows.append({
                        'ROLLNO': rollno,
                        'COLUMN': col,
                        'Scanned': row1[col],
                        'Extracted': row2[col],
                        'Image Path': image_path
                    })
                else:
                    discrepancies.append({
                        'ROLLNO': rollno,
                        'COLUMN': col,
                        'Scanned': row1[col],
                        'Extracted': row2[col],
                        'Image Path': image_path
                    })

    discrepancy_df = pd.DataFrame(discrepancies)
    qpseries_df = pd.DataFrame(qpseries_rows)

    rollno_counts = discrepancy_df['ROLLNO'].value_counts().reset_index()
    rollno_counts.columns = ['ROLLNO', 'Count_in_Discrepancies']

    total_discrepancies = len(discrepancies)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Discrepancies'

    headers = ['ROLLNO', 'COLUMN', 'Scanned', 'Extracted', 'Image Path']
    for col_num, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row_num, discrepancy in enumerate(discrepancies, start=2):
        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_num, column=col_num, value=discrepancy.get(header, ''))
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write total discrepancies count clearly
    total_discrepancies_label = f"Total Discrepancies: {total_discrepancies}"
    total_discrepancies_cell = f'A{len(discrepancies) + 3}'  # Place it a few rows below the discrepancies
    total_discrepancies_cell_obj = worksheet[total_discrepancies_cell]
    total_discrepancies_cell_obj.value = total_discrepancies_label
    total_discrepancies_cell_obj.font = Font(bold=True, color="FF0000")  # Red color for visibility
    total_discrepancies_cell_obj.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow background

    for row_num, (rollno, count) in enumerate(rollno_counts.values, start=len(discrepancies) + 5):
        worksheet[f'A{row_num}'] = rollno
        worksheet[f'B{row_num}'] = count

    qpseries_sheet = workbook.create_sheet(title='QPSERIES Count')
    qpseries_headers = ['ROLLNO', 'COLUMN', 'Scanned', 'Extracted', 'Image Path']
    for col_num, header in enumerate(qpseries_headers, start=1):
        cell = qpseries_sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row_num, qpseries_row in enumerate(qpseries_rows, start=2):
        for col_num, value in enumerate(qpseries_row.values(), start=1):
            cell = qpseries_sheet.cell(row=row_num, column=col_num, value=value)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    workbook.save(output_path)

    if include_qpseries == 'y' and 'QPSERIES' in columns_to_compare:
        mb.showinfo("QPSERIES Included", "QPSERIES column was included in comparison. See 'QPSERIES Count' sheet for details.", parent=root)
    else:
        mb.showinfo("Success", "Data exported successfully to the chosen path!", parent=root)

if __name__ == "__main__":
    upload_files()
